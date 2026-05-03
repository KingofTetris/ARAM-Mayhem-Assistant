import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HF = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="1A237E")
HA = Alignment(horizontal="center", vertical="center", wrap_text=True)
SF = Font(name="Microsoft YaHei", bold=True, size=11)
SFILL = PatternFill("solid", fgColor="E8EAF6")
NF = Font(name="Microsoft YaHei", size=10)
NA = Alignment(horizontal="center", vertical="center", wrap_text=True)
LA = Alignment(horizontal="left", vertical="center", wrap_text=True)
TB = Border(left=Side("thin","BDBDBD"),right=Side("thin","BDBDBD"),top=Side("thin","BDBDBD"),bottom=Side("thin","BDBDBD"))
CF = PatternFill("solid", fgColor="C8E6C9")
IF = PatternFill("solid", fgColor="FFF9C4")
NF2 = PatternFill("solid", fgColor="FFECB3")
AF = PatternFill("solid", fgColor="E3F2FD")
BF = PatternFill("solid", fgColor="F3E5F5")

def hdr(ws, hs, ws_list):
    for i,(h,w) in enumerate(zip(hs,ws_list),1):
        c=ws.cell(1,i,h); c.font=HF; c.fill=HFILL; c.alignment=HA; c.border=TB
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[1].height=36

def fill(ws, data, sc=1, stc=None, pc=None, lc=None, rh=28):
    if lc is None: lc=[]
    for ri,rd in enumerate(data,2):
        for ci,v in enumerate(rd,1):
            c=ws.cell(ri,ci,v); c.font=NF; c.alignment=LA if ci in lc else NA; c.border=TB
            if ci==sc: c.font=SF; c.fill=SFILL
            if stc and ci==stc:
                if v=="已完成": c.fill=CF
                elif v=="进行中": c.fill=IF
                elif v=="未开始": c.fill=NF2
            if pc and ci==pc:
                c.number_format='0%'; c.value=v/100.0
        ws.row_dimensions[ri].height=rh

# ============================================================
# Sheet 1: 总体概览
# ============================================================
ws1 = wb.active; ws1.title="总体概览"
hdr(ws1, ["指标","安卓端","后端","整体"], [28,40,40,16])
overview = [
    ["项目名称","海克斯大乱斗信息差助手 (ARAM Mayhem Assistant)","海克斯大乱斗信息差助手 (ARAM Mayhem Assistant)","—"],
    ["项目路径","D:\\androidProjects\\ARAM_Mayhem_Assistant","D:\\ideaProjects\\aram-server","—"],
    ["技术栈","Java 21 + XML + ViewBinding + Hilt","Java 21 + Spring Boot 3.3.5 + MyBatis Plus","—"],
    ["源文件数","8 个手写 Java + 11 模块 build.gradle","34 个 Java 源文件 + pom.xml + 2 yml","42+"],
    ["构建状态","Gradle BUILD SUCCESS (284 tasks)","Maven BUILD SUCCESS (34 files compiled)","✓"],
    ["APK/JAR 产出","app-debug.apk (7.8 MB)","aram-server-1.0.0.jar","✓"],
    ["当前阶段","M3 进行中 (核心模块完成,8组件待开发)","M2 已完成 (数据库+API+认证全部就绪)","—"],
    ["整体完成率","55%","100% (M1+M2)","72%"],
    ["下一里程碑","M3: 通用Android组件开发","M4: 英雄模块后端API","—"],
    ["关键阻塞项","8个通用UI组件未开发","DataInitializer种子数据待补充","—"],
]
for ri,rd in enumerate(overview,2):
    for ci,v in enumerate(rd,1):
        c=ws1.cell(ri,ci,v); c.font=NF; c.alignment=LA if ci!=4 else NA; c.border=TB
        if ci==1: c.font=SF; c.fill=SFILL
        if ci==2: c.fill=AF
        if ci==3: c.fill=BF
    ws1.row_dimensions[ri].height=28

# ============================================================
# Sheet 2: 安卓端进度
# ============================================================
ws2 = wb.create_sheet("安卓端进度")

# Section A: 已完成工作
ws2.cell(1,1,"A. 已完成工作").font=Font(name="Microsoft YaHei",bold=True,size=12,color="1A237E")
ws2.merge_cells("A1:H1")
ws2.cell(1,1).fill=PatternFill("solid",fgColor="E8EAF6")
ws2.cell(1,1).alignment=LA; ws2.cell(1,1).border=TB
ws2.row_dimensions[1].height=32

hdr_a = ["序号","工作项","具体内容","涉及模块","完成时间","状态","完成度","验证结果"]
widths_a = [6,28,48,18,20,10,10,28]
for i,(h,w) in enumerate(zip(hdr_a,widths_a),1):
    c=ws2.cell(2,i,h); c.font=HF; c.fill=PatternFill("solid",fgColor="1565C0"); c.alignment=HA; c.border=TB
    ws2.column_dimensions[get_column_letter(i)].width=w
ws2.row_dimensions[2].height=30

android_done = [
    [1,"项目脚手架搭建","创建Java+XML项目(com.aram.mayhem,JDK21),11个Gradle模块","app+core×4+feature×5+nav","2026-05-03 11:30","已完成",100,"settings.gradle+11个build.gradle"],
    [2,"Gradle多模块配置","Groovy DSL,统一依赖版本管理(AGP 8.5.2, Hilt 2.51.1等)","全部模块","2026-05-03 17:45","已完成",100,"ext{}统一版本号"],
    [3,"ViewBinding+Hilt配置","app+5个feature模块启用viewBinding, Hilt @AndroidEntryPoint","app+feature×5","2026-05-03 17:45","已完成",100,"MayhemApplication+6个Fragment"],
    [4,"Navigation导航图","5 Tab底部导航(英雄/符文/玩法/公告/我的),navigation_graph.xml","app","2026-05-03 17:30","已完成",100,"BottomNavigationView+NavController"],
    [5,"core-common公共模块","Constants(BASE_URL+API路径), Result<T>, Tier枚举(S+/S/A/B/C)","core-common","2026-05-03 17:30","已完成",100,"3个Java类"],
    [6,"core-ui共享资源体系","25色值+35字符串+30尺寸+14样式+13个drawable","core-ui","2026-05-03 18:00","已完成",100,"colors/strings/dimens/themes/drawable"],
    [7,"5个Fragment占位","HeroList/AugmentList/CommunityFeed/BulletinList/Profile","feature×5","2026-05-03 17:30","已完成",100,"@AndroidEntryPoint+ViewBinding"],
    [8,"网络安全配置","network_security_config.xml允许局域网明文HTTP","app","2026-05-03 17:45","已完成",100,"192.168.1.100+localhost"],
    [9,"启动图标","ic_launcher_foreground/background adaptive icon","core-ui+app","2026-05-03 18:00","已完成",100,"mipmap-anydpi-v26"],
    [10,"ProGuard混淆规则","keep Retrofit/Room/Gson/Hilt类","app","2026-05-03 17:45","已完成",100,"proguard-rules.pro"],
    [11,"资源修复(关键)","9个AndroidManifest+共享资源迁移+导航图迁移+图标修复","全部模块","2026-05-03 18:00","已完成",100,"Gradle BUILD SUCCESS"],
]
for ri,rd in enumerate(android_done,3):
    for ci,v in enumerate(rd,1):
        c=ws2.cell(ri,ci,v); c.font=NF; c.alignment=LA if ci in [3,8] else NA; c.border=TB
        if ci==6 and v=="已完成": c.fill=CF
        if ci==7: c.number_format='0%'; c.value=v/100.0
    ws2.row_dimensions[ri].height=28

# Section B: 待完成任务
row_b = 3 + len(android_done) + 1
ws2.cell(row_b,1,"B. 待完成任务").font=Font(name="Microsoft YaHei",bold=True,size=12,color="E65100")
ws2.merge_cells(f"A{row_b}:H{row_b}")
ws2.cell(row_b,1).fill=PatternFill("solid",fgColor="FFF3E0")
ws2.cell(row_b,1).alignment=LA; ws2.cell(row_b,1).border=TB
ws2.row_dimensions[row_b].height=32

for i,(h,w) in enumerate(zip(["序号","任务项","具体内容","涉及模块","目标完成时间","优先级","预估工时","依赖关系"],widths_a),1):
    c=ws2.cell(row_b+1,i,h); c.font=HF; c.fill=PatternFill("solid",fgColor="E65100"); c.alignment=HA; c.border=TB
ws2.row_dimensions[row_b+1].height=30

android_todo = [
    [1,"AuthInterceptor","OkHttp Interceptor自动注入Authorization Bearer token","core-network","2026-05-04","高","2h","JWT认证完成后端"],
    [2,"TokenRefreshInterceptor","401响应自动刷新Token+重放原始请求","core-network","2026-05-04","高","3h","AuthInterceptor"],
    [3,"Retrofit API接口定义","HeroApi/AugmentApi/CommunityApi/BulletinApi/UserApi","core-network","2026-05-04","高","4h","AuthInterceptor"],
    [4,"Room Database实现","AppDatabase+HeroDao/AugmentDao/StrategyDao","core-data","2026-05-04","高","4h","Entity定义"],
    [5,"EncryptedSharedPreferences","Token安全存储(accessToken+refreshToken)","core-data","2026-05-04","高","2h","无"],
    [6,"HeroCardAdapter","RecyclerView适配器:头像+名称+TierBadge+胜率,12dp圆角CardView","feature-hero","2026-05-05","高","4h","TierBadgeView"],
    [7,"AugmentCardAdapter","品质色边框(棱彩=渐变金紫/金=金黄/银=银灰)","feature-augment","2026-05-05","高","3h","QualityChip"],
    [8,"TierBadgeView","自定义View: S+/S/A/B/C五色标签","core-ui","2026-05-05","高","3h","无"],
    [9,"SearchToolbar","TextInputLayout+TextWatcher防抖300ms","core-ui","2026-05-05","中","2h","无"],
    [10,"PaginationScrollListener","RecyclerView.OnScrollListener上滑加载更多","core-ui","2026-05-05","中","2h","无"],
    [11,"StatefulLayout","Loading/Empty/Error/Content四状态切换","core-ui","2026-05-05","中","3h","无"],
    [12,"QualityChip","品质Chip(棱彩=紫金/金=金色/银=银色)","core-ui","2026-05-05","中","2h","无"],
    [13,"BalanceBar","Canvas自绘平衡修正条形图(绿/红/灰+数值)","core-ui","2026-05-05","中","3h","无"],
    [14,"HeroRepository","先Room缓存→网络请求→更新Room→LiveData推送","feature-hero","2026-05-06","高","4h","Retrofit+Room"],
    [15,"HeroListFragment完整实现","RecyclerView按Tier分组(StickyHeader)+筛选+搜索+离线","feature-hero","2026-05-07","高","6h","HeroCardAdapter+Repository"],
]
for ri,rd in enumerate(android_todo,row_b+2):
    for ci,v in enumerate(rd,1):
        c=ws2.cell(ri,ci,v); c.font=NF; c.alignment=LA if ci in [3,8] else NA; c.border=TB
        if ci==5: c.fill=IF
        if ci==6:
            if v=="高": c.fill=PatternFill("solid",fgColor="FFCDD2")
            elif v=="中": c.fill=PatternFill("solid",fgColor="FFF9C4")
    ws2.row_dimensions[ri].height=28

# Section C: 阶段性计划
row_c = row_b + 2 + len(android_todo) + 1
ws2.cell(row_c,1,"C. 阶段性计划与目标时间节点").font=Font(name="Microsoft YaHei",bold=True,size=12,color="2E7D32")
ws2.merge_cells(f"A{row_c}:H{row_c}")
ws2.cell(row_c,1).fill=PatternFill("solid",fgColor="E8F5E9")
ws2.cell(row_c,1).alignment=LA; ws2.cell(row_c,1).border=TB
ws2.row_dimensions[row_c].height=32

for i,(h,w) in enumerate(zip(["阶段","目标","关键交付物","目标完成时间","当前状态","完成率","阻塞项","备注"],widths_a),1):
    c=ws2.cell(row_c+1,i,h); c.font=HF; c.fill=PatternFill("solid",fgColor="2E7D32"); c.alignment=HA; c.border=TB
ws2.row_dimensions[row_c+1].height=30

android_plan = [
    ["M3","Android UI框架搭建","8个通用组件+Retrofit接口+Room实现","2026-05-05","进行中",55,"8个UI组件未开发","核心模块已完成"],
    ["M4","英雄模块","HeroListFragment+HeroDetailFragment+HeroRepository","2026-05-09","未开始",0,"依赖M3组件完成","后端API已就绪"],
    ["M5","强化符文模块","AugmentListFragment+SynergyProgress+Recommend","2026-05-13","未开始",0,"依赖M3+M4","后端API待开发"],
    ["M6","社区模块","CommunityFeedFragment+PublishStrategy+VoteButton","2026-05-17","未开始",0,"依赖M3+认证","后端API待开发"],
    ["M7","公告模块","BulletinCarousel+BulletinList+VersionTrapBanner","2026-05-19","未开始",0,"依赖M3","后端API待开发"],
    ["M8","个人中心","ProfileFragment+SettingsFragment+MyStrategies","2026-05-21","未开始",0,"依赖认证模块","后端API待开发"],
    ["M10","测试与打包","Espresso测试+ProGuard+Release APK","2026-05-29","未开始",0,"依赖所有功能模块","LeakCanary集成"],
    ["M11","本地部署验证","ADB install+真机验证+离线测试","2026-05-31","未开始",0,"依赖M10","10项验证清单"],
]
for ri,rd in enumerate(android_plan,row_c+2):
    for ci,v in enumerate(rd,1):
        c=ws2.cell(ri,ci,v); c.font=NF; c.alignment=LA if ci in [3,7,8] else NA; c.border=TB
        if ci==1: c.font=SF; c.fill=SFILL
        if ci==5:
            if v=="进行中": c.fill=IF
            elif v=="未开始": c.fill=NF2
        if ci==6: c.number_format='0%'; c.value=v/100.0
    ws2.row_dimensions[ri].height=28

# ============================================================
# Sheet 3: 后端进度
# ============================================================
ws3 = wb.create_sheet("后端进度")

ws3.cell(1,1,"A. 已完成工作").font=Font(name="Microsoft YaHei",bold=True,size=12,color="4A148C")
ws3.merge_cells("A1:H1")
ws3.cell(1,1).fill=PatternFill("solid",fgColor="F3E5F5")
ws3.cell(1,1).alignment=LA; ws3.cell(1,1).border=TB
ws3.row_dimensions[1].height=32

for i,(h,w) in enumerate(zip(["序号","工作项","具体内容","涉及包","完成时间","状态","完成度","验证结果"],widths_a),1):
    c=ws3.cell(2,i,h); c.font=HF; c.fill=PatternFill("solid",fgColor="6A1B9A"); c.alignment=HA; c.border=TB
    ws3.column_dimensions[get_column_letter(i)].width=w
ws3.row_dimensions[2].height=30

backend_done = [
    [1,"Spring Boot项目初始化","Spring Boot 3.3.5 + Maven + Java 21 项目结构","根包","2026-05-03 11:30","已完成",100,"AramServerApplication.java"],
    [2,"Maven pom.xml配置","MyBatis Plus 3.5.9 + Spring Security + jjwt 0.12.6 + SpringDoc","根","2026-05-03 17:45","已完成",100,"34个依赖"],
    [3,"application.yml配置","MySQL(127.0.0.1:3306/aram_mayhem)+Redis+JWT+MyBatis Plus","resources","2026-05-03 17:45","已完成",100,"utf8mb4+Jackson序列化"],
    [4,"MySQL数据库建模","9张表DDL+10个索引,utf8mb4_unicode_ci","database","2026-05-03 20:30","已完成",100,"tb_user/hero/hero_modifier/augment/strategy等"],
    [5,"9个Entity类","@TableName+@TableId+@TableField,逻辑删除deleted字段","entity","2026-05-03 20:30","已完成",100,"User/Hero/HeroModifier/Augment/Strategy/Vote/Bulletin等"],
    [6,"9个Mapper接口","BaseMapper<T>,继承MyBatis Plus通用CRUD","mapper","2026-05-03 20:30","已完成",100,"UserMapper/HeroMapper/AugmentMapper等"],
    [7,"统一响应体Result<T>","code+message+data+timestamp,success()/error()静态工厂","common","2026-05-03 20:45","已完成",100,"@JsonInclude(NON_NULL)"],
    [8,"全局异常处理器","@RestControllerAdvice,处理Validation/Business/Unknown异常","common","2026-05-03 20:45","已完成",100,"GlobalExceptionHandler"],
    [9,"BusinessException","自定义业务异常(code+message)","common","2026-05-03 20:45","已完成",100,"400默认码"],
    [10,"CORS跨域配置","CorsFilter,允许所有Origin+标准方法+凭证","config","2026-05-03 20:45","已完成",100,"CorsConfig.java"],
    [11,"Redis缓存配置","RedisTemplate+Jackson2JsonRedisSerializer,支持Java8时间","config","2026-05-03 20:45","已完成",100,"RedisConfig.java"],
    [12,"SpringDoc OpenAPI","Swagger UI /swagger-ui.html,api-docs /v3/api-docs","config","2026-05-03 20:45","已完成",100,"springdoc 2.6.0"],
    [13,"JwtTokenProvider","HMAC-SHA密钥签名,Access 15min+Refresh 7d","security","2026-05-03 21:30","已完成",100,"jjwt 0.12.6生成/解析/验证"],
    [14,"JwtAuthenticationFilter","OncePerRequestFilter,Bearer Token提取+验证+SecurityContext","security","2026-05-03 21:30","已完成",100,"注入UserDetailsService"],
    [15,"SecurityFilterChain","CSRF禁用+STATELESS会话+/api/auth/**放行+其余需认证","config","2026-05-03 21:30","已完成",100,"SecurityConfig.java"],
    [16,"CustomUserDetailsService","UserDetailsService实现,email查询+权限加载","security","2026-05-03 21:30","已完成",100,"BCrypt密码验证"],
    [17,"注册接口","POST /api/auth/register,邮箱唯一+BCrypt加密","controller+service","2026-05-03 21:30","已完成",100,"RegisterRequest DTO"],
    [18,"登录接口","POST /api/auth/login,凭证验证→Access+Refresh Token","controller+service","2026-05-03 21:30","已完成",100,"AuthResponse DTO"],
    [19,"Token刷新接口","POST /api/auth/refresh,RefreshToken续期","controller+service","2026-05-03 21:30","已完成",100,"RefreshTokenRequest DTO"],
    [20,"4个DTO类","RegisterRequest/LoginRequest/RefreshTokenRequest/AuthResponse","dto","2026-05-03 21:30","已完成",100,"@Valid+@NotBlank校验"],
    [21,"AuthService业务层","register/login/refresh完整业务逻辑","service","2026-05-03 21:30","已完成",100,"邮箱唯一校验+密码匹配+Token生成"],
    [22,"start-server.bat","一键mvn package+java -jar启动脚本","根","2026-05-03 17:45","已完成",100,"--spring.profiles.active=local"],
]
for ri,rd in enumerate(backend_done,3):
    for ci,v in enumerate(rd,1):
        c=ws3.cell(ri,ci,v); c.font=NF; c.alignment=LA if ci in [3,8] else NA; c.border=TB
        if ci==6 and v=="已完成": c.fill=CF
        if ci==7: c.number_format='0%'; c.value=v/100.0
    ws3.row_dimensions[ri].height=28

# Section B: 待完成任务
row_b2 = 3 + len(backend_done) + 1
ws3.cell(row_b2,1,"B. 待完成任务").font=Font(name="Microsoft YaHei",bold=True,size=12,color="E65100")
ws3.merge_cells(f"A{row_b2}:H{row_b2}")
ws3.cell(row_b2,1).fill=PatternFill("solid",fgColor="FFF3E0")
ws3.cell(row_b2,1).alignment=LA; ws3.cell(row_b2,1).border=TB
ws3.row_dimensions[row_b2].height=32

for i,(h,w) in enumerate(zip(["序号","任务项","具体内容","涉及包","目标完成时间","优先级","预估工时","依赖关系"],widths_a),1):
    c=ws3.cell(row_b2+1,i,h); c.font=HF; c.fill=PatternFill("solid",fgColor="E65100"); c.alignment=HA; c.border=TB
ws3.row_dimensions[row_b2+1].height=30

backend_todo = [
    [1,"DataInitializer种子数据","启动时预置英雄+强化符文基础数据","config","2026-05-06","高","3h","数据库已建表"],
    [2,"HeroController+HeroService","GET /api/heroes分页查询+GET /api/heroes/{id}详情","controller+service","2026-05-06","高","4h","Entity+Mapper已就绪"],
    [3,"HeroService缓存策略","Redis缓存英雄列表10min+详情30min","service","2026-05-06","高","2h","RedisConfig已就绪"],
    [4,"AugmentController+AugmentService","GET /api/augments品质+套装筛选+Redis 30min缓存","controller+service","2026-05-10","高","4h","Entity+Mapper已就绪"],
    [5,"Augment推荐接口","POST /api/augments/recommend,评分排序","controller+service","2026-05-12","中","4h","Augment基础CRUD"],
    [6,"StrategyController+StrategyService","GET /api/strategies hot/latest分页+POST发布","controller+service","2026-05-14","高","6h","认证+Entity已就绪"],
    [7,"VoteController","POST /api/strategies/{id}/vote upsert+计数更新","controller+service","2026-05-16","中","3h","Strategy基础CRUD"],
    [8,"BulletinController","GET /api/bulletins分页+首页最新3条","controller+service","2026-05-18","中","3h","Entity+Mapper已就绪"],
    [9,"AdminController","PUT /api/admin/heroes/{id}/trap-mark管理员接口","controller+service","2026-05-18","低","2h","Security角色校验"],
    [10,"UserController","GET /api/users/me/profile+PATCH更新","controller+service","2026-05-20","中","3h","认证已就绪"],
    [11,"RiotDataDragonClient","RestTemplate调用Data Dragon API解析英雄JSON","service","2026-05-22","中","4h","HeroService"],
    [12,"AramDataCollector","Jsoup解析U.GG/ARAMMayhem页面数据","service","2026-05-22","中","6h","无"],
    [13,"DataSyncScheduler","@Scheduled每6小时全量同步+缓存预热","service","2026-05-24","中","4h","DataCollector"],
    [14,"backup-mysql.bat","mysqldump自动备份脚本","根","2026-05-30","低","1h","无"],
]
for ri,rd in enumerate(backend_todo,row_b2+2):
    for ci,v in enumerate(rd,1):
        c=ws3.cell(ri,ci,v); c.font=NF; c.alignment=LA if ci in [3,8] else NA; c.border=TB
        if ci==5: c.fill=IF
        if ci==6:
            if v=="高": c.fill=PatternFill("solid",fgColor="FFCDD2")
            elif v=="中": c.fill=PatternFill("solid",fgColor="FFF9C4")
            elif v=="低": c.fill=PatternFill("solid",fgColor="C8E6C9")
    ws3.row_dimensions[ri].height=28

# Section C: 阶段性计划
row_c2 = row_b2 + 2 + len(backend_todo) + 1
ws3.cell(row_c2,1,"C. 阶段性计划与目标时间节点").font=Font(name="Microsoft YaHei",bold=True,size=12,color="2E7D32")
ws3.merge_cells(f"A{row_c2}:H{row_c2}")
ws3.cell(row_c2,1).fill=PatternFill("solid",fgColor="E8F5E9")
ws3.cell(row_c2,1).alignment=LA; ws3.cell(row_c2,1).border=TB
ws3.row_dimensions[row_c2].height=32

for i,(h,w) in enumerate(zip(["阶段","目标","关键交付物","目标完成时间","当前状态","完成率","阻塞项","备注"],widths_a),1):
    c=ws3.cell(row_c2+1,i,h); c.font=HF; c.fill=PatternFill("solid",fgColor="2E7D32"); c.alignment=HA; c.border=TB
ws3.row_dimensions[row_c2+1].height=30

backend_plan = [
    ["M1","项目初始化","Spring Boot项目+Maven+Git","2026-05-03","已完成",100,"无","已完成"],
    ["M2","基础架构","9表+索引+Entity/Mapper+RESTful框架+JWT认证","2026-05-03","已完成",100,"无","34源文件编译通过"],
    ["M4","英雄模块API","HeroController+HeroService+Redis缓存+种子数据","2026-05-07","未开始",0,"DataInitializer待开发","Entity/Mapper已就绪"],
    ["M5","强化符文API","AugmentController+推荐算法+套装进度","2026-05-11","未开始",0,"依赖M4","Entity/Mapper已就绪"],
    ["M6","社区模块API","StrategyController+VoteController+玩法发布","2026-05-15","未开始",0,"依赖M4+认证","认证已就绪"],
    ["M7","公告+管理API","BulletinController+AdminController+陷阱标记","2026-05-19","未开始",0,"依赖M4","Entity已就绪"],
    ["M8","个人中心API","UserController+资料更新+我的投稿","2026-05-21","未开始",0,"依赖认证","认证已就绪"],
    ["M9","数据管线","DataDragon+AramDataCollector+6h同步+缓存预热","2026-05-24","未开始",0,"依赖M4+M5","外部API依赖"],
    ["M10","测试","JUnit5+Mockito+MockMvc集成测试","2026-05-27","未开始",0,"依赖所有API","覆盖率目标80%"],
    ["M11","本地部署验证","start-server.bat+健康检查+备份脚本","2026-05-31","未开始",0,"依赖M10","10项验证清单"],
]
for ri,rd in enumerate(backend_plan,row_c2+2):
    for ci,v in enumerate(rd,1):
        c=ws3.cell(ri,ci,v); c.font=NF; c.alignment=LA if ci in [3,7,8] else NA; c.border=TB
        if ci==1: c.font=SF; c.fill=SFILL
        if ci==5:
            if v=="已完成": c.fill=CF
            elif v=="未开始": c.fill=NF2
        if ci==6: c.number_format='0%'; c.value=v/100.0
    ws3.row_dimensions[ri].height=28

# ============================================================
# Sheet 4: 更新日志
# ============================================================
ws4 = wb.create_sheet("更新日志")
hdr(ws4, ["更新时间","更新内容","操作端","操作人","影响范围"], [22,60,10,10,20])
log_data = [
    ["2026-05-03 09:00:00","项目初始化开始：创建Android+后端项目结构","双端","全栈开发","M1"],
    ["2026-05-03 11:30:00","Android: 11模块项目脚手架搭建完成","安卓端","Android开发","M1-Task1"],
    ["2026-05-03 11:30:00","后端: Spring Boot 3.3.5+Maven项目创建完成","后端","后端开发","M1-Task1"],
    ["2026-05-03 17:45:00","Android: Gradle多模块+ViewBinding+Hilt+网络安全配置完成","安卓端","Android开发","M1-Task2"],
    ["2026-05-03 17:45:00","后端: Maven pom.xml+application.yml配置完成","后端","后端开发","M1-Task2"],
    ["2026-05-03 20:30:00","后端: MySQL 9表+10索引+9 Entity+9 Mapper全部完成","后端","后端开发","M2-Task3"],
    ["2026-05-03 20:45:00","后端: Result+全局异常+CORS+Redis+SpringDoc框架完成","后端","后端开发","M2-Task4"],
    ["2026-05-03 21:30:00","后端: JWT认证全链路完成(注册/登录/刷新+Filter+Security)","后端","后端开发","M2-Task5"],
    ["2026-05-03 17:30:00","Android: 核心模块搭建+5个Fragment占位+导航图完成","安卓端","Android开发","M3-Task6"],
    ["2026-05-03 18:00:00","Android: 资源修复(9个AndroidManifest+共享资源迁移+图标修复)","安卓端","Android开发","M3-Task6"],
    ["2026-05-03 18:05:00","Android: Gradle BUILD SUCCESS, app-debug.apk 7.8MB产出","安卓端","Android开发","全项目"],
    ["2026-05-03 18:10:00","后端: Maven BUILD SUCCESS, 34源文件编译通过","后端","后端开发","全项目"],
]
for ri,rd in enumerate(log_data,2):
    for ci,v in enumerate(rd,1):
        c=ws4.cell(ri,ci,v); c.font=NF; c.alignment=LA if ci==2 else NA; c.border=TB
        if ci==3:
            if v=="安卓端": c.fill=AF
            elif v=="后端": c.fill=BF
            elif v=="双端": c.fill=SFILL
    ws4.row_dimensions[ri].height=24

# ============================================================
# Sheet 5: 里程碑概览
# ============================================================
ws5 = wb.create_sheet("里程碑概览")
hdr(ws5, ["里程碑","名称","安卓端状态","安卓端完成率","后端状态","后端完成率","整体完成率","关键交付物"], [10,20,12,12,12,12,12,50])
ms_data = [
    ["M1","项目初始化","已完成",1.0,"已完成",1.0,1.0,"Android 11模块+后端Spring Boot+Gradle/Maven+Git"],
    ["M2","基础架构","已完成",1.0,"已完成",1.0,1.0,"Android资源体系+后端9表+API框架+JWT认证"],
    ["M3","Android UI框架","进行中",0.55,"—","—",0.55,"8个通用组件+Retrofit+Room(仅安卓端)"],
    ["M4","英雄模块","未开始",0.0,"未开始",0.0,0.0,"HeroListFragment+HeroDetailFragment+HeroController"],
    ["M5","强化符文模块","未开始",0.0,"未开始",0.0,0.0,"AugmentListFragment+AugmentController+推荐算法"],
    ["M6","社区模块","未开始",0.0,"未开始",0.0,0.0,"CommunityFeedFragment+StrategyController+Vote"],
    ["M7","版本与公告","未开始",0.0,"未开始",0.0,0.0,"BulletinCarousel+BulletinController+陷阱标记"],
    ["M8","个人中心","未开始",0.0,"未开始",0.0,0.0,"ProfileFragment+UserController+Settings"],
    ["M9","数据管线","—","—","未开始",0.0,0.0,"DataDragon+AramDataCollector+6h同步(仅后端)"],
    ["M10","测试与打包","未开始",0.0,"未开始",0.0,0.0,"Espresso+MockMvc+ProGuard+Release APK"],
    ["M11","本地部署验证","未开始",0.0,"未开始",0.0,0.0,"ADB install+真机验证+backup-mysql.bat"],
]
for ri,rd in enumerate(ms_data,2):
    for ci,v in enumerate(rd,1):
        c=ws5.cell(ri,ci,v); c.font=NF; c.alignment=LA if ci==8 else NA; c.border=TB
        if ci==1: c.font=SF; c.fill=SFILL
        if ci==3 or ci==5:
            if v=="已完成": c.fill=CF
            elif v=="进行中": c.fill=IF
            elif v=="未开始": c.fill=NF2
        if ci in [4,6,7]: c.number_format='0%'
    ws5.row_dimensions[ri].height=28

# Save to 3 locations
paths = [
    r"D:\androidProjects\ARAM_Mayhem_Assistant\project_progress_tracker.xlsx",
    r"D:\ideaProjects\aram-server\project_progress_tracker.xlsx",
    r"D:\traeProjects\ARAM_Mayhem_Assistant\project_progress_tracker.xlsx",
]
for p in paths:
    wb.save(p)
    print(f"Saved: {p}")
